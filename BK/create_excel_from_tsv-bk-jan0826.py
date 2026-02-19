#!/usr/bin/env python3 -u
"""
Script: create_excel_from_tsv.py
Purpose: Convert TSV query results to Excel file with multiple sheets
         Automatically splits large datasets into multiple sheets if they exceed Excel's row limit
         Uses unbuffered output for real-time progress reporting
Author: Infrastructure Team
Usage: python3 create_excel_from_tsv.py <tsv_directory> <output_excel_file> <server_name>
"""

import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

# Force unbuffered output
sys.stdout.reconfigure(line_buffering=True) if hasattr(sys.stdout, 'reconfigure') else None

# Excel row limit (1,048,576 rows total, reserve 1 for header)
EXCEL_MAX_ROWS = 1048575  # Maximum data rows per sheet

def create_excel_from_tsv_files(tsv_dir, excel_file, server_name):
    """
    Create an Excel file with multiple sheets from TSV files
    Automatically splits sheets if data exceeds Excel's row limit
    
    Args:
        tsv_dir: Directory containing TSV files
        excel_file: Output Excel file path
        server_name: Name of the server (for Server column)
    """
    
    # Abbreviated sheet names to avoid 31-character Excel limit
    # Maps full query name to abbreviated sheet name
    sheet_name_mapping = {
        "Group_Hierarchy_Path": "GroupHierarchyPath",
        "Gateway_Count_By_Group": "GatewayCountByGroup",
        "User_List_With_Groups": "UserListWithGroups"
    }
    
    # Expected column headers for each query (when no data)
    # These match the SELECT columns from each query's SQL
    expected_columns = {
        "Group_Hierarchy_Path": ["rootid", "rootgroup", "groupid", "groupname"],
        "Gateway_Count_By_Group": ["groupname", "groupid", "statlabel", "value", "num_gateways"],
        "User_List_With_Groups": ["groupid", "groupname", "username", "email"]
    }
    
    # Map TSV filenames (without extension) to query names for lookup
    tsv_to_query_map = {
        "Group_Hierarchy_Path": "Group_Hierarchy_Path",
        "Gateway_Count_By_Group": "Gateway_Count_By_Group", 
        "User_List_With_Groups": "User_List_With_Groups"
    }
    
    # Query definitions for "Queries" index sheet
    queries = {
        "Group_Hierarchy_Path": """WITH RECURSIVE grp_path (groupid, label, rootid, rootlabel) AS
(
  SELECT groupid, label, groupid AS rootid, label AS rootlabel
  FROM im_group WHERE parentgroupid = 0
  UNION ALL
  SELECT c.groupid, c.label, sup.rootid, sup.rootlabel
  FROM grp_path AS sup
  JOIN im_group c ON sup.groupid = c.parentgroupid
)
SELECT rootid, rootlabel AS rootgroup, groupid, label AS groupname
FROM grp_path
ORDER BY rootid, groupid;""",
        
        "Gateway_Count_By_Group": """SELECT grp.label AS groupname, grp.groupid, stat.statlabel, latest.value, count(*) AS num_gateways
FROM im_group grp
  JOIN im_node node ON grp.groupid = node.nodeid
  JOIN im_lateststatitem latest ON latest.nodeid = node.nodeid
  JOIN im_stat stat ON latest.statid = stat.statid
WHERE grp.groupid IS NOT NULL
AND node.groupid <> 0
GROUP BY grp.groupid, grp.label, stat.statlabel, latest.value
ORDER BY grp.label, grp.groupid, stat.statlabel, latest.value;""",

        "User_List_With_Groups": """SELECT perm.groupid, g.label AS groupname, u.loginname AS username, u.email
FROM im_user u
  JOIN im_permissiongroup perm ON u.userid = perm.userid
  JOIN im_group g ON perm.groupid = g.groupid
ORDER BY perm.groupid, g.label;"""
    }
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Create "Queries" index sheet
    print(f"Creating 'Queries' index sheet...")
    queries_sheet = wb.create_sheet("Queries", 0)
    queries_sheet['A1'] = "Query Name"
    queries_sheet['B1'] = "SQL Query"
    queries_sheet['C1'] = "Server"
    queries_sheet['D1'] = "Generated Date"
    
    for cell in ['A1', 'B1', 'C1', 'D1']:
        queries_sheet[cell].font = header_font
        queries_sheet[cell].fill = header_fill
        queries_sheet[cell].alignment = header_alignment
    
    row = 2
    for query_name, query_sql in queries.items():
        queries_sheet[f'A{row}'] = query_name
        queries_sheet[f'B{row}'] = query_sql
        queries_sheet[f'C{row}'] = server_name
        queries_sheet[f'D{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        queries_sheet[f'B{row}'].alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    
    queries_sheet.column_dimensions['A'].width = 30
    queries_sheet.column_dimensions['B'].width = 80
    queries_sheet.column_dimensions['C'].width = 15
    queries_sheet.column_dimensions['D'].width = 20
    
    print(f"'Queries' sheet created")
    sys.stdout.flush()  # Flush output immediately
    
    # Process each TSV file
    sheets_created = 0
    
    for query_name in queries.keys():
        tsv_file = os.path.join(tsv_dir, f"{query_name}.tsv")
        
        # Get abbreviated sheet name (max 31 chars for Excel)
        base_sheet_name = sheet_name_mapping.get(query_name, query_name[:20])
        
        if not os.path.exists(tsv_file):
            print(f"TSV not found: {tsv_file} - creating placeholder sheet with expected columns")
            print(f"  Query name: {query_name}")
            print(f"  Base sheet name: {base_sheet_name}")
            
            ws = wb.create_sheet(base_sheet_name)
            
            # Get expected column headers for this query
            expected_cols = expected_columns.get(query_name, [])
            print(f"  Expected columns for '{query_name}': {expected_cols}")
            print(f"  Number of expected columns: {len(expected_cols)}")
            
            if len(expected_cols) == 0:
                print(f"  WARNING: No expected columns found for query '{query_name}'!")
                print(f"  Available keys in expected_columns: {list(expected_columns.keys())}")
            
            # Write Server column header
            ws.cell(row=1, column=1, value="Server")
            ws.cell(row=1, column=1).font = header_font
            ws.cell(row=1, column=1).fill = header_fill
            ws.cell(row=1, column=1).alignment = header_alignment
            
            # Write all expected column headers
            for col_idx, col_name in enumerate(expected_cols, start=2):
                print(f"  Writing column {col_idx}: {col_name}")
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Write data row: server name + "No Data Available" for each column
            ws.cell(row=2, column=1, value=server_name)
            for col_idx in range(2, len(expected_cols) + 2):
                ws.cell(row=2, column=col_idx, value="No Data Available")
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 15
            for col_idx in range(2, len(expected_cols) + 2):
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                ws.column_dimensions[col_letter].width = 20
            
            print(f"  Created placeholder sheet '{base_sheet_name}' with {len(expected_cols)} data columns (+ Server column)")
            sheets_created += 1
            continue
        
        print(f"Processing: {query_name}")
        sys.stdout.flush()  # Flush immediately
        
        try:
            # Read all lines from TSV
            with open(tsv_file, 'r') as f:
                lines = [line.rstrip('\n') for line in f.readlines()]
            
            # Check if this is a "No Data Available" placeholder file from shell script
            if len(lines) == 1 and lines[0].strip() == "No Data Available":
                print(f"  TSV contains 'No Data Available' placeholder - creating proper sheet")
                ws = wb.create_sheet(base_sheet_name)
                
                # Use expected headers for this query
                headers = expected_columns.get(query_name, [])
                
                # Write Server column header
                ws.cell(row=1, column=1, value="Server")
                ws.cell(row=1, column=1).font = header_font
                ws.cell(row=1, column=1).fill = header_fill
                ws.cell(row=1, column=1).alignment = header_alignment
                
                # Write all other column headers
                for col_idx, header in enumerate(headers, start=2):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Write data row: server name + "No Data Available" for each column
                ws.cell(row=2, column=1, value=server_name)
                for col_idx in range(2, len(headers) + 2):
                    ws.cell(row=2, column=col_idx, value="No Data Available")
                
                # Adjust column widths
                ws.column_dimensions['A'].width = 15
                for col_idx in range(2, len(headers) + 2):
                    col_letter = ws.cell(row=1, column=col_idx).column_letter
                    ws.column_dimensions[col_letter].width = 20
                
                print(f"  Created placeholder with {len(headers)} columns")
                sheets_created += 1
                continue
            
            if len(lines) == 0 or len(lines) == 1:
                # Empty file or only header
                print(f"  Empty or header-only TSV - creating placeholder with expected columns")
                ws = wb.create_sheet(base_sheet_name)
                
                # Determine headers - use from file if available, else use expected
                if len(lines) == 1:
                    headers = lines[0].split('\t')
                else:
                    headers = expected_columns.get(query_name, [])
                
                # Write Server column header
                ws.cell(row=1, column=1, value="Server")
                ws.cell(row=1, column=1).font = header_font
                ws.cell(row=1, column=1).fill = header_fill
                ws.cell(row=1, column=1).alignment = header_alignment
                
                # Write all other column headers
                for col_idx, header in enumerate(headers, start=2):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Write data row: server name + "No Data Available" for each column
                ws.cell(row=2, column=1, value=server_name)
                for col_idx in range(2, len(headers) + 2):
                    ws.cell(row=2, column=col_idx, value="No Data Available")
                
                # Adjust column widths
                ws.column_dimensions['A'].width = 15
                for col_idx in range(2, len(headers) + 2):
                    col_letter = ws.cell(row=1, column=col_idx).column_letter
                    ws.column_dimensions[col_letter].width = 20
                
                print(f"  Created placeholder with {len(headers)} columns")
                sheets_created += 1
                continue
            
            # Extract headers and data
            headers = lines[0].split('\t')
            data_lines = lines[1:]
            total_rows = len(data_lines)
            
            print(f"  Total data rows: {total_rows:,}")
            
            # Determine if dataset is large (optimize processing for large datasets)
            is_large_dataset = total_rows > 100000
            
            if is_large_dataset:
                print(f"  Large dataset detected - using optimized processing")
            
            # Check if we need to split
            if total_rows > EXCEL_MAX_ROWS:
                # Calculate number of sheets needed
                num_sheets = (total_rows + EXCEL_MAX_ROWS - 1) // EXCEL_MAX_ROWS
                print(f"  Splitting into {num_sheets} sheets (exceeds {EXCEL_MAX_ROWS:,} row limit)")
                sys.stdout.flush()  # Flush immediately
                
                # Create multiple sheets
                for sheet_idx in range(num_sheets):
                    # Use abbreviated name with part number: GatewayCount_1, GatewayCount_2, etc.
                    sheet_name = f"{base_sheet_name}_{sheet_idx + 1}"
                    
                    # Ensure sheet name is still under 31 characters
                    if len(sheet_name) > 31:
                        sheet_name = f"{base_sheet_name[:25]}_{sheet_idx + 1}"
                    
                    start_row = sheet_idx * EXCEL_MAX_ROWS
                    end_row = min(start_row + EXCEL_MAX_ROWS, total_rows)
                    chunk = data_lines[start_row:end_row]
                    
                    print(f"  Creating '{sheet_name}': rows {start_row+1:,} to {end_row:,} ({len(chunk):,} rows)")
                    sys.stdout.flush()  # Flush immediately
                    
                    ws = wb.create_sheet(sheet_name)
                    
                    # Write headers with Server column first
                    ws.cell(row=1, column=1, value="Server")
                    ws.cell(row=1, column=1).font = header_font
                    ws.cell(row=1, column=1).fill = header_fill
                    ws.cell(row=1, column=1).alignment = header_alignment
                    
                    for col_idx, header in enumerate(headers, start=2):
                        cell = ws.cell(row=1, column=col_idx, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    
                    # Write data rows
                    for row_idx, line in enumerate(chunk, start=2):
                        # Show progress for large datasets
                        if is_large_dataset and row_idx % 100000 == 0:
                            print(f"    Progress: {row_idx-1:,}/{len(chunk):,} rows written")
                            sys.stdout.flush()  # Flush progress updates
                        
                        ws.cell(row=row_idx, column=1, value=server_name)
                        
                        values = line.split('\t')
                        for col_idx, value in enumerate(values, start=2):
                            # Skip type conversion for large datasets to save CPU
                            if is_large_dataset:
                                ws.cell(row=row_idx, column=col_idx, value=value)
                            else:
                                try:
                                    if value.isdigit():
                                        ws.cell(row=row_idx, column=col_idx, value=int(value))
                                    elif '.' in value and value.replace('.', '', 1).isdigit():
                                        ws.cell(row=row_idx, column=col_idx, value=float(value))
                                    else:
                                        ws.cell(row=row_idx, column=col_idx, value=value)
                                except:
                                    ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Adjust column widths - simplified for large datasets
                    if is_large_dataset:
                        # Use fixed widths for large datasets to save CPU time
                        ws.column_dimensions['A'].width = 15  # Server column
                        for col_idx in range(2, len(headers) + 2):
                            col_letter = ws.cell(row=1, column=col_idx).column_letter
                            ws.column_dimensions[col_letter].width = 20  # Fixed width
                    else:
                        # Sample first 100 rows for column width
                        sample_rows = min(100, len(chunk))
                        for col in range(1, len(headers) + 2):
                            max_len = 0
                            for row in range(1, sample_rows + 2):
                                try:
                                    val = ws.cell(row=row, column=col).value
                                    if val and len(str(val)) > max_len:
                                        max_len = len(str(val))
                                except:
                                    pass
                            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 2, 50)
                    
                    sheets_created += 1
                
                print(f"  Successfully split {query_name} into {num_sheets} sheets")
            
            else:
                # Single sheet - data fits
                print(f"  Creating single sheet (fits within limit)")
                ws = wb.create_sheet(base_sheet_name)
                
                # Write headers
                ws.cell(row=1, column=1, value="Server")
                ws.cell(row=1, column=1).font = header_font
                ws.cell(row=1, column=1).fill = header_fill
                ws.cell(row=1, column=1).alignment = header_alignment
                
                for col_idx, header in enumerate(headers, start=2):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Write data
                for row_idx, line in enumerate(data_lines, start=2):
                    # Show progress for large datasets
                    if is_large_dataset and row_idx % 100000 == 0:
                        print(f"    Progress: {row_idx-1:,}/{total_rows:,} rows written")
                    
                    ws.cell(row=row_idx, column=1, value=server_name)
                    
                    values = line.split('\t')
                    for col_idx, value in enumerate(values, start=2):
                        # Skip type conversion for large datasets to save CPU
                        if is_large_dataset:
                            ws.cell(row=row_idx, column=col_idx, value=value)
                        else:
                            try:
                                if value.isdigit():
                                    ws.cell(row=row_idx, column=col_idx, value=int(value))
                                elif '.' in value and value.replace('.', '', 1).isdigit():
                                    ws.cell(row=row_idx, column=col_idx, value=float(value))
                                else:
                                    ws.cell(row=row_idx, column=col_idx, value=value)
                            except:
                                ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Adjust columns - simplified for large datasets
                if is_large_dataset:
                    # Fixed widths to save CPU
                    ws.column_dimensions['A'].width = 15
                    for col_idx in range(2, len(headers) + 2):
                        col_letter = ws.cell(row=1, column=col_idx).column_letter
                        ws.column_dimensions[col_letter].width = 20
                else:
                    # Auto-adjust for small datasets
                    for column in ws.columns:
                        max_len = 0
                        col_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if cell.value and len(str(cell.value)) > max_len:
                                    max_len = len(str(cell.value))
                            except:
                                pass
                        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
                
                sheets_created += 1
                print(f"  Sheet '{query_name}' created: {total_rows:,} rows")
        
        except Exception as e:
            print(f"ERROR processing {query_name}: {str(e)}")
            ws = wb.create_sheet(query_name)
            ws['A1'] = "Error"
            ws['A1'].font = Font(bold=True, color="FF0000")
            ws['A2'] = f"Failed: {str(e)}"
            sheets_created += 1
    
    # Save workbook
    try:
        wb.save(excel_file)
        print(f"SUCCESS: Excel created with {sheets_created + 1} total sheets")
        sys.stdout.flush()
        return 0
    except Exception as e:
        print(f"ERROR saving Excel: {str(e)}")
        sys.stdout.flush()
        return 1

if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("Usage: python3 create_excel_from_tsv.py <tsv_dir> <output_excel> <server_name>")
        print("")
        print("Example:")
        print("  python3 create_excel_from_tsv.py /tmp/tsv_AMM01 /tmp/AMM01_20251010_120000.xlsx AMM01")
        sys.exit(1)
    
    tsv_directory = sys.argv[1]
    output_excel = sys.argv[2]
    server_name = sys.argv[3]
    
    if not os.path.isdir(tsv_directory):
        print(f"ERROR: TSV directory not found: {tsv_directory}")
        sys.exit(1)
    
    exit_code = create_excel_from_tsv_files(tsv_directory, output_excel, server_name)
    sys.exit(exit_code)