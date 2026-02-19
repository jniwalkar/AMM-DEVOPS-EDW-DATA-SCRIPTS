#!/usr/bin/env python3
"""
Script: create_excel_from_tsv.py
Purpose: Convert TSV query results to Excel file with multiple sheets
Author: Infrastructure Team
Usage: python3 create_excel_from_tsv.py <tsv_directory> <output_excel_file> <server_name>
"""

import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

def create_excel_from_tsv_files(tsv_dir, excel_file, server_name):
    """
    Create an Excel file with multiple sheets from TSV files

    Args:
        tsv_dir: Directory containing TSV files
        excel_file: Output Excel file path
        server_name: Name of the server (for metadata and Server column)
    """

    # Query definitions for "Queries" sheet
    queries = {
        "Group_Hierarchy_Path": """WITH RECURSIVE grp_path (groupid, label, rootid, rootlabel) AS
(
  SELECT groupid, label, groupid AS rootid, label AS rootlabel
  FROM im_group WHERE parentgroupid = 0
  UNION ALL
  SELECT c.groupid, c.label, sup.rootid, sup.rootlabel
  FROM grp_path AS sup
  JOIN im_group c ON sup.groupid = c.parentgroupid
)
SELECT rootid, rootlabel AS rootgroup, groupid, label AS groupname
FROM grp_path
ORDER BY rootid, groupid;""",

        "Gateway_Count_By_Group": """SELECT grp.label AS groupname, grp.groupid, stat.statlabel, latest.value, count(*) AS num_gateways
FROM im_group grp
  JOIN im_node node ON grp.groupid = node.groupid
  JOIN im_lateststatitem latest ON latest.nodeid = node.nodeid
  JOIN im_stat stat ON latest.statid = stat.statid
WHERE grp.groupid IS NOT NULL
AND node.groupid <> 0
GROUP BY grp.groupid, grp.label, stat.statlabel, latest.value
ORDER BY grp.label, grp.groupid, stat.statlabel, latest.value;""",

        "User_List_With_Groups": """SELECT perm.groupid, g.label AS groupname, u.loginname AS username, u.email
FROM im_user u
  JOIN im_permissiongroup perm ON u.userid = perm.userid
  JOIN im_group g ON perm.groupid = g.groupid
ORDER BY perm.groupid, g.label;"""
    }

    # Create Excel workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Create "Queries" index sheet (Sheet 1)
    print(f"Creating 'Queries' index sheet...")
    queries_sheet = wb.create_sheet("Queries", 0)
    queries_sheet['A1'] = "Query Name"
    queries_sheet['B1'] = "SQL Query"
    queries_sheet['C1'] = "Server"
    queries_sheet['D1'] = "Generated Date"

    # Apply header styling
    for cell in ['A1', 'B1', 'C1', 'D1']:
        queries_sheet[cell].font = header_font
        queries_sheet[cell].fill = header_fill
        queries_sheet[cell].alignment = header_alignment

    # Add queries to index sheet
    row = 2
    for query_name, query_sql in queries.items():
        queries_sheet[f'A{row}'] = query_name
        queries_sheet[f'B{row}'] = query_sql
        queries_sheet[f'C{row}'] = server_name
        queries_sheet[f'D{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        queries_sheet[f'B{row}'].alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    # Adjust column widths
    queries_sheet.column_dimensions['A'].width = 30
    queries_sheet.column_dimensions['B'].width = 80
    queries_sheet.column_dimensions['C'].width = 15
    queries_sheet.column_dimensions['D'].width = 20

    print(f"'Queries' sheet created with {len(queries)} query definitions")

    # Process each TSV file and create data sheets
    sheets_created = 0
    for query_name in queries.keys():
        tsv_file = os.path.join(tsv_dir, f"{query_name}.tsv")

        # Always create sheet, even if TSV doesn't exist
        ws = wb.create_sheet(query_name)

        if os.path.exists(tsv_file):
            print(f"Processing TSV file: {tsv_file}")

            try:
                with open(tsv_file, 'r') as f:
                    lines = [line.rstrip('\n') for line in f.readlines()]

                if len(lines) > 0:
                    # First line is headers from MySQL
                    headers = lines[0].split('\t')

                    # Add "Server" as first column header
                    ws.cell(row=1, column=1, value="Server")
                    ws.cell(row=1, column=1).font = header_font
                    ws.cell(row=1, column=1).fill = header_fill
                    ws.cell(row=1, column=1).alignment = header_alignment

                    # Write original headers starting from column 2
                    for col_idx, header in enumerate(headers, start=2):
                        cell = ws.cell(row=1, column=col_idx, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment

                    # Write data rows with server name in first column
                    for row_idx, line in enumerate(lines[1:], start=2):
                        # Add server name in first column
                        ws.cell(row=row_idx, column=1, value=server_name)

                        # Add query data starting from column 2
                        values = line.split('\t')
                        for col_idx, value in enumerate(values, start=2):
                            # Try to convert to number if possible
                            try:
                                if value.isdigit():
                                    ws.cell(row=row_idx, column=col_idx, value=int(value))
                                elif '.' in value and value.replace('.', '', 1).isdigit():
                                    ws.cell(row=row_idx, column=col_idx, value=float(value))
                                else:
                                    ws.cell(row=row_idx, column=col_idx, value=value)
                            except:
                                ws.cell(row=row_idx, column=col_idx, value=value)

                    # Auto-adjust column widths
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if cell.value and len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width

                    sheets_created += 1
                    print(f"Sheet '{query_name}' created: {len(lines)-1} data rows")
                else:
                    # Empty TSV - create sheet with "No Data" message
                    print(f"TSV file '{query_name}' is empty - creating empty sheet")
                    ws['A1'] = "Server"
                    ws['A1'].font = header_font
                    ws['A1'].fill = header_fill
                    ws['A1'].alignment = header_alignment
                    ws['A2'] = server_name
                    ws['B1'] = "No Data Available"
                    ws['B1'].font = Font(bold=True, color="FF6600")
                    sheets_created += 1

            except Exception as e:
                print(f"Error processing TSV file '{query_name}': {str(e)}")
                ws['A1'] = "Server"
                ws['A1'].font = header_font
                ws['A1'].fill = header_fill
                ws['A2'] = server_name
                ws['B1'] = "Error"
                ws['B1'].font = Font(bold=True, color="FF0000")
                ws['B2'] = f"Failed to process data: {str(e)}"
                sheets_created += 1
        else:
            # TSV file doesn't exist - still create sheet with placeholder
            print(f"TSV file not found: {tsv_file} - creating empty sheet")
            ws['A1'] = "Server"
            ws['A1'].font = header_font
            ws['A1'].fill = header_fill
            ws['A1'].alignment = header_alignment
            ws['A2'] = server_name
            ws['B1'] = "No Data Available"
            ws['B1'].font = Font(bold=True, color="FF6600")
            ws['B2'] = "No data returned from query"
            sheets_created += 1

    # Save Excel file
    try:
        wb.save(excel_file)
        print(f"SUCCESS: Excel file created: {excel_file}")
        print(f"Total sheets: {sheets_created + 1} (1 index + {sheets_created} data sheets)")
        return 0
    except Exception as e:
        print(f"ERROR: Failed to save Excel file: {str(e)}")
        return 1

if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("Usage: python3 create_excel_from_tsv.py <tsv_directory> <output_excel_file> <server_name>")
        print("")
        print("Example:")
        print("  python3 create_excel_from_tsv.py /tmp/tsv_AMM01 /tmp/AMM01_20251009_120000.xlsx AMM01")
        sys.exit(1)

    tsv_directory = sys.argv[1]
    output_excel = sys.argv[2]
    server_name = sys.argv[3]

    if not os.path.isdir(tsv_directory):
        print(f"ERROR: TSV directory not found: {tsv_directory}")
        sys.exit(1)

    # Create Excel file
    exit_code = create_excel_from_tsv_files(tsv_directory, output_excel, server_name)
    sys.exit(exit_code)
