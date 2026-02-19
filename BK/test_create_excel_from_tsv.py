#!/usr/bin/env python3
"""
Test script for create_excel_from_tsv.py
Tests various scenarios including empty data and failure cases
"""

import os
import sys
import tempfile
import shutil
from openpyxl import load_workbook

# Import the main function
sys.path.insert(0, os.path.dirname(__file__))
from create_excel_from_tsv import create_excel_from_tsv_files

def print_test_header(test_name):
    """Print a formatted test header"""
    print("\n" + "="*80)
    print(f"TEST: {test_name}")
    print("="*80)

def verify_sheet_structure(ws, expected_cols, server_name):
    """Verify a sheet has the correct structure"""
    print(f"  Verifying sheet '{ws.title}'...")
    
    # Check header row
    actual_headers = []
    for col_idx in range(1, len(expected_cols) + 2):
        cell_value = ws.cell(row=1, column=col_idx).value
        actual_headers.append(cell_value)
    
    expected_headers = ["Server"] + expected_cols
    
    print(f"    Expected headers: {expected_headers}")
    print(f"    Actual headers:   {actual_headers}")
    
    if actual_headers != expected_headers:
        print(f"    ❌ FAIL: Headers don't match!")
        return False
    
    # Check data row
    server_cell = ws.cell(row=2, column=1).value
    print(f"    Server value: {server_cell}")
    
    if server_cell != server_name:
        print(f"    ❌ FAIL: Server name doesn't match! Expected '{server_name}', got '{server_cell}'")
        return False
    
    # Check "No Data Available" in other columns
    for col_idx in range(2, len(expected_cols) + 2):
        cell_value = ws.cell(row=2, column=col_idx).value
        if cell_value != "No Data Available":
            print(f"    ❌ FAIL: Column {col_idx} should be 'No Data Available', got '{cell_value}'")
            return False
    
    print(f"    ✅ PASS: Sheet structure is correct!")
    return True

def test_scenario_1_missing_tsv_files():
    """Test when TSV files don't exist"""
    print_test_header("Scenario 1: Missing TSV Files")
    
    temp_dir = tempfile.mkdtemp()
    tsv_dir = os.path.join(temp_dir, "tsv")
    os.makedirs(tsv_dir)
    
    output_excel = os.path.join(temp_dir, "test_missing.xlsx")
    server_name = "TEST01"
    
    print(f"TSV Directory: {tsv_dir}")
    print(f"Output Excel: {output_excel}")
    print("Creating Excel with NO TSV files...")
    
    # Run the function
    result = create_excel_from_tsv_files(tsv_dir, output_excel, server_name)
    
    if result != 0:
        print("❌ FAIL: Function returned non-zero exit code")
        shutil.rmtree(temp_dir)
        return False
    
    # Verify Excel file
    print("\nVerifying Excel file...")
    wb = load_workbook(output_excel)
    
    expected_sheets = ["Queries", "GroupHierarchyPath", "GatewayCountByGroup", "UserListWithGroups"]
    actual_sheets = wb.sheetnames
    
    print(f"Expected sheets: {expected_sheets}")
    print(f"Actual sheets:   {actual_sheets}")
    
    if actual_sheets != expected_sheets:
        print("❌ FAIL: Sheet names don't match!")
        shutil.rmtree(temp_dir)
        return False
    
    # Verify each data sheet
    expected_columns = {
        "GroupHierarchyPath": ["rootid", "rootgroup", "groupid", "groupname"],
        "GatewayCountByGroup": ["groupname", "groupid", "statlabel", "value", "num_gateways"],
        "UserListWithGroups": ["groupid", "groupname", "username", "email"]
    }
    
    all_passed = True
    for sheet_name, cols in expected_columns.items():
        ws = wb[sheet_name]
        if not verify_sheet_structure(ws, cols, server_name):
            all_passed = False
    
    shutil.rmtree(temp_dir)
    
    if all_passed:
        print("\n✅ TEST PASSED: All missing TSV files handled correctly")
        return True
    else:
        print("\n❌ TEST FAILED: Some sheets have incorrect structure")
        return False

def test_scenario_2_empty_tsv_files():
    """Test when TSV files exist but are empty"""
    print_test_header("Scenario 2: Empty TSV Files")
    
    temp_dir = tempfile.mkdtemp()
    tsv_dir = os.path.join(temp_dir, "tsv")
    os.makedirs(tsv_dir)
    
    # Create empty TSV files
    queries = ["Group_Hierarchy_Path", "Gateway_Count_By_Group", "User_List_With_Groups"]
    for query in queries:
        tsv_file = os.path.join(tsv_dir, f"{query}.tsv")
        open(tsv_file, 'w').close()  # Create empty file
        print(f"Created empty file: {tsv_file}")
    
    output_excel = os.path.join(temp_dir, "test_empty.xlsx")
    server_name = "TEST02"
    
    print(f"\nOutput Excel: {output_excel}")
    print("Creating Excel with EMPTY TSV files...")
    
    # Run the function
    result = create_excel_from_tsv_files(tsv_dir, output_excel, server_name)
    
    if result != 0:
        print("❌ FAIL: Function returned non-zero exit code")
        shutil.rmtree(temp_dir)
        return False
    
    # Verify Excel file
    print("\nVerifying Excel file...")
    wb = load_workbook(output_excel)
    
    expected_columns = {
        "GroupHierarchyPath": ["rootid", "rootgroup", "groupid", "groupname"],
        "GatewayCountByGroup": ["groupname", "groupid", "statlabel", "value", "num_gateways"],
        "UserListWithGroups": ["groupid", "groupname", "username", "email"]
    }
    
    all_passed = True
    for sheet_name, cols in expected_columns.items():
        ws = wb[sheet_name]
        if not verify_sheet_structure(ws, cols, server_name):
            all_passed = False
    
    shutil.rmtree(temp_dir)
    
    if all_passed:
        print("\n✅ TEST PASSED: All empty TSV files handled correctly")
        return True
    else:
        print("\n❌ TEST FAILED: Some sheets have incorrect structure")
        return False

def test_scenario_3_header_only_tsv():
    """Test when TSV files have only headers, no data"""
    print_test_header("Scenario 3: TSV Files with Headers Only")
    
    temp_dir = tempfile.mkdtemp()
    tsv_dir = os.path.join(temp_dir, "tsv")
    os.makedirs(tsv_dir)
    
    # Create TSV files with headers only
    tsv_data = {
        "Group_Hierarchy_Path": "rootid\trootgroup\tgroupid\tgroupname\n",
        "Gateway_Count_By_Group": "groupname\tgroupid\tstatlabel\tvalue\tnum_gateways\n",
        "User_List_With_Groups": "groupid\tgroupname\tusername\temail\n"
    }
    
    for query, header_line in tsv_data.items():
        tsv_file = os.path.join(tsv_dir, f"{query}.tsv")
        with open(tsv_file, 'w') as f:
            f.write(header_line)
        print(f"Created header-only file: {tsv_file}")
    
    output_excel = os.path.join(temp_dir, "test_header_only.xlsx")
    server_name = "TEST03"
    
    print(f"\nOutput Excel: {output_excel}")
    print("Creating Excel with HEADER-ONLY TSV files...")
    
    # Run the function
    result = create_excel_from_tsv_files(tsv_dir, output_excel, server_name)
    
    if result != 0:
        print("❌ FAIL: Function returned non-zero exit code")
        shutil.rmtree(temp_dir)
        return False
    
    # Verify Excel file
    print("\nVerifying Excel file...")
    wb = load_workbook(output_excel)
    
    expected_columns = {
        "GroupHierarchyPath": ["rootid", "rootgroup", "groupid", "groupname"],
        "GatewayCountByGroup": ["groupname", "groupid", "statlabel", "value", "num_gateways"],
        "UserListWithGroups": ["groupid", "groupname", "username", "email"]
    }
    
    all_passed = True
    for sheet_name, cols in expected_columns.items():
        ws = wb[sheet_name]
        if not verify_sheet_structure(ws, cols, server_name):
            all_passed = False
    
    shutil.rmtree(temp_dir)
    
    if all_passed:
        print("\n✅ TEST PASSED: All header-only TSV files handled correctly")
        return True
    else:
        print("\n❌ TEST FAILED: Some sheets have incorrect structure")
        return False

def test_scenario_4_valid_data():
    """Test with valid data"""
    print_test_header("Scenario 4: Valid Data in TSV Files")
    
    temp_dir = tempfile.mkdtemp()
    tsv_dir = os.path.join(temp_dir, "tsv")
    os.makedirs(tsv_dir)
    
    # Create TSV files with valid data
    tsv_data = {
        "Group_Hierarchy_Path": "rootid\trootgroup\tgroupid\tgroupname\n1\tRoot1\t10\tGroup1\n2\tRoot2\t20\tGroup2\n",
        "Gateway_Count_By_Group": "groupname\tgroupid\tstatlabel\tvalue\tnum_gateways\nGroup1\t1\tOnline\t1\t5\nGroup2\t2\tOffline\t0\t3\n",
        "User_List_With_Groups": "groupid\tgroupname\tusername\temail\n1\tAdmins\tjdoe\tjdoe@example.com\n2\tUsers\tjsmith\tjsmith@example.com\n"
    }
    
    for query, content in tsv_data.items():
        tsv_file = os.path.join(tsv_dir, f"{query}.tsv")
        with open(tsv_file, 'w') as f:
            f.write(content)
        print(f"Created valid data file: {tsv_file}")
    
    output_excel = os.path.join(temp_dir, "test_valid.xlsx")
    server_name = "TEST04"
    
    print(f"\nOutput Excel: {output_excel}")
    print("Creating Excel with VALID TSV data...")
    
    # Run the function
    result = create_excel_from_tsv_files(tsv_dir, output_excel, server_name)
    
    if result != 0:
        print("❌ FAIL: Function returned non-zero exit code")
        shutil.rmtree(temp_dir)
        return False
    
    # Verify Excel file
    print("\nVerifying Excel file...")
    wb = load_workbook(output_excel)
    
    # Check UserListWithGroups sheet has actual data
    ws = wb["UserListWithGroups"]
    print(f"\nChecking UserListWithGroups sheet:")
    print(f"  Row 1: {[ws.cell(1, c).value for c in range(1, 6)]}")
    print(f"  Row 2: {[ws.cell(2, c).value for c in range(1, 6)]}")
    print(f"  Row 3: {[ws.cell(3, c).value for c in range(1, 6)]}")
    
    # Verify we have actual data, not "No Data Available"
    if ws.cell(2, 2).value == "No Data Available":
        print("❌ FAIL: Valid data was treated as empty!")
        shutil.rmtree(temp_dir)
        return False
    
    if ws.cell(2, 2).value != 1 and ws.cell(2, 2).value != "1":
        print(f"❌ FAIL: Expected groupid '1', got '{ws.cell(2, 2).value}'")
        shutil.rmtree(temp_dir)
        return False
    
    shutil.rmtree(temp_dir)
    print("\n✅ TEST PASSED: Valid data processed correctly")
    return True

def test_scenario_5_mixed_conditions():
    """Test with mixed conditions: some files missing, some empty, some with data"""
    print_test_header("Scenario 5: Mixed Conditions")
    
    temp_dir = tempfile.mkdtemp()
    tsv_dir = os.path.join(temp_dir, "tsv")
    os.makedirs(tsv_dir)
    
    # Group_Hierarchy_Path: Missing (don't create)
    # Gateway_Count_By_Group: Empty file
    tsv_file = os.path.join(tsv_dir, "Gateway_Count_By_Group.tsv")
    open(tsv_file, 'w').close()
    print(f"Created empty file: {tsv_file}")
    
    # User_List_With_Groups: Valid data
    tsv_file = os.path.join(tsv_dir, "User_List_With_Groups.tsv")
    with open(tsv_file, 'w') as f:
        f.write("groupid\tgroupname\tusername\temail\n1\tAdmins\tjdoe\tjdoe@example.com\n")
    print(f"Created valid data file: {tsv_file}")
    
    output_excel = os.path.join(temp_dir, "test_mixed.xlsx")
    server_name = "TEST05"
    
    print(f"\nOutput Excel: {output_excel}")
    print("Creating Excel with MIXED conditions...")
    
    # Run the function
    result = create_excel_from_tsv_files(tsv_dir, output_excel, server_name)
    
    if result != 0:
        print("❌ FAIL: Function returned non-zero exit code")
        shutil.rmtree(temp_dir)
        return False
    
    # Verify Excel file
    print("\nVerifying Excel file...")
    wb = load_workbook(output_excel)
    
    # Check GroupHierarchyPath (missing file)
    ws = wb["GroupHierarchyPath"]
    print(f"\nGroupHierarchyPath (missing file):")
    if not verify_sheet_structure(ws, ["rootid", "rootgroup", "groupid", "groupname"], server_name):
        shutil.rmtree(temp_dir)
        return False
    
    # Check GatewayCountByGroup (empty file)
    ws = wb["GatewayCountByGroup"]
    print(f"\nGatewayCountByGroup (empty file):")
    if not verify_sheet_structure(ws, ["groupname", "groupid", "statlabel", "value", "num_gateways"], server_name):
        shutil.rmtree(temp_dir)
        return False
    
    # Check UserListWithGroups (valid data)
    ws = wb["UserListWithGroups"]
    print(f"\nUserListWithGroups (valid data):")
    print(f"  Row 1: {[ws.cell(1, c).value for c in range(1, 6)]}")
    print(f"  Row 2: {[ws.cell(2, c).value for c in range(1, 6)]}")
    
    if ws.cell(2, 2).value == "No Data Available":
        print("❌ FAIL: Valid data was treated as empty!")
        shutil.rmtree(temp_dir)
        return False
    
    shutil.rmtree(temp_dir)
    print("\n✅ TEST PASSED: Mixed conditions handled correctly")
    return True

def run_all_tests():
    """Run all test scenarios"""
    print("\n" + "#"*80)
    print("# EXCEL TSV CONVERTER TEST SUITE")
    print("#"*80)
    
    tests = [
        test_scenario_1_missing_tsv_files,
        test_scenario_2_empty_tsv_files,
        test_scenario_3_header_only_tsv,
        test_scenario_4_valid_data,
        test_scenario_5_mixed_conditions
    ]
    
    results = []
    for test_func in tests:
        try:
            result = test_func()
            results.append((test_func.__name__, result))
        except Exception as e:
            print(f"\n❌ EXCEPTION in {test_func.__name__}: {str(e)}")
            import traceback
            traceback.print_exc()
            results.append((test_func.__name__, False))
    
    # Summary
    print("\n" + "#"*80)
    print("# TEST SUMMARY")
    print("#"*80)
    
    passed = sum(1 for _, result in results if result)
    total = len(results)
    
    for test_name, result in results:
        status = "✅ PASS" if result else "❌ FAIL"
        print(f"{status}: {test_name}")
    
    print(f"\nTotal: {passed}/{total} tests passed")
    
    if passed == total:
        print("\n🎉 ALL TESTS PASSED! 🎉")
        return 0
    else:
        print(f"\n⚠️  {total - passed} test(s) failed")
        return 1

if __name__ == "__main__":
    sys.exit(run_all_tests())